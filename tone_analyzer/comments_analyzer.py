
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import json
from os.path import join, dirname
from watson_developer_cloud import ToneAnalyzerV3

tone_analyzer = ToneAnalyzerV3(
    username='',
    password='',
    version='2017-09-26')

comments_file = "./Comments_2017.xlsx"

wb = load_workbook(comments_file)
ws = wb.active


def CreateHeaders(ws):
     tonetypes = ['Anger','Fear','Joy','Sadness','Analytical','Confident','Tentative']
     new_header_range = ws['Y1':'AE1'][0]
     headerfill = PatternFill("solid", bgColor='FFC0C0C0', fgColor='FFC0C0C0')
     for cell in range(len(new_header_range)):
             new_header_range[cell].value = tonetypes[cell]
             new_header_range[cell].fill = headerfill

def LoadSheet(file):
     wb = load_workbook(file)
     ws = wb.active
     return ws

def AnalyzeTone(ws, tone_analyzer):
	tonetype_dict = {"Anger":"Y","Fear":"Z","Joy":"AA","Sadness":"AB","Analytical":"AC","Confident":"AD","Tentative":"AE"}
	for cellObj in ws['J2':'J'+str(ws.max_row)]:
		row = cellObj[0].row
		tonereturn = tone_analyzer.tone(tone_input=cellObj[0].value,content_type="text/plain", sentences=False)
		for tone in tonereturn['document_tone']['tones']:
			column = tonetype_dict[tone["tone_name"]]
			ws[column+str(row)] = tone["score"]
		print(row)



CreateHeaders(ws)
AnalyzeTone(ws, tone_analyzer)
wb.save(filename="Comments_2017_tone.xlsx")


